"""
    V1 기록
    1. 프로토타입 완성.
    2. 영상 등록일 저장 요청 있음.
"""
from apiclient.discovery import build
from apiclient.errors import HttpError
from oauth2client.tools import argparser
from openpyxl import Workbook

import json
import urllib

_keyword = str(input())

wb = Workbook()
ws = wb.create_sheet(_keyword)
"""
    주의!
    1. DEVELOPER_KEY 절대 외부에 노출하지 마세요.
"""
DEVELOPER_KEY = "AIzaSyCJEHbjf_SBS7dkswshNZRlnhEGxnF7nKo"
YOUTUBE_API_SERVICE_NAME = "youtube"
YOUTUBE_API_VERSION = "v3"
FREEBASE_SEARCH_URL = "https://www.googleapis.com/freebase/v1/search?%s"

youtube = build(YOUTUBE_API_SERVICE_NAME,
                YOUTUBE_API_VERSION,
                developerKey=DEVELOPER_KEY)

next_page_token = None
search_video_counter = 1

_row = 1
ws.cell(column=1, row=_row, value="번호")
ws.cell(column=2, row=_row, value="채널이름")
ws.cell(column=3, row=_row, value="동영상이름")
ws.cell(column=4, row=_row, value="비디오ID(유튜브 검색 시 참고)")
ws.cell(column=5, row=_row, value="동영상게시날짜")
ws.cell(column=6, row=_row, value="조회수")
ws.cell(column=7, row=_row, value="좋아요수")
ws.cell(column=8, row=_row, value="댓글수")
ws.cell(column=9, row=_row, value="길이")

for search_counter in range(1, 6):
    if next_page_token:
        search_response = youtube.search().list(q=_keyword,
                                                part="snippet",
                                                pageToken=next_page_token,
                                                order='viewCount',
                                                maxResults=50).execute()
    else:
        search_response = youtube.search().list(q=_keyword,
                                                part="snippet",
                                                order='viewCount',
                                                maxResults=50).execute()

    for item in search_response['items']:
        _row += 1
        try:
            channel_title = item['snippet']['channelTitle']
        except:
            channel_title = '채널 이름 없음.'

        try:
            title = item['snippet']['title']
        except:
            title = '동영상 제목 없음.'

        try:
            videoId = item['id']['videoId']
        except:
            videoId = None

        try:
            published_at = item['snippet']['publishedAt']
        except:
            published_at = None

        ws.cell(column=1, row=_row, value=search_video_counter)
        ws.cell(column=2, row=_row, value=channel_title)
        ws.cell(column=3, row=_row, value=title)
        ws.cell(column=4, row=_row, value=videoId)
        ws.cell(column=5, row=_row, value=published_at)
        print('#', search_video_counter, '\t', '채널 이름\t', channel_title, '\t',
              '동영상 제목\t', title, '\t', '동영상 id\t', videoId)

        if videoId:
            video_search_response_1 = youtube.videos().list(
                part="statistics", id=videoId).execute()
            
            video_search_response_2 = youtube.videos().list(
                part="contentDetails", id=videoId).execute()

            try:
                view_count = video_search_response_1['items'][0]['statistics'][
                    'viewCount']
            except:
                view_count = 0
            try:
                like_count = video_search_response_1['items'][0]['statistics'][
                    'likeCount'] or 0
            except:
                like_count = 0
            try:
                comment_count = video_search_response_1['items'][0][
                    'statistics']['commentCount'] or 0
            except:
                comment_count = 0
            try:
                duration = video_search_response_2['items'][0]['contentDetails']['duration']
            except:
                duration = 0

            ws.cell(column=6, row=_row, value=view_count)
            ws.cell(column=7, row=_row, value=like_count)
            ws.cell(column=8, row=_row, value=comment_count)
            ws.cell(column=9, row=_row, value=duration)
            print('뷰\t', view_count, '\t', '좋아요\t', like_count, '\t', '댓글\t',
                  comment_count, '길이\t', duration)

        search_video_counter += 1

    if search_response['nextPageToken']:
        next_page_token = search_response['nextPageToken']
    else:
        print('토큰이 없다.')
        break

wb.save(filename=_keyword + ".xlsx")